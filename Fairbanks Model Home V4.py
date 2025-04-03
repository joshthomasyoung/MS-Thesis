# Model Fairbanks Home heating and cooling 1 year simulation
# Version 4 (Commercial equipment COP and Heating capacity data added, conservative performance calculations added, added supplemental boiler logic, added optimization loop, changed RSC solar collector and tank sizes to optimized sizes)
# any term denoted by Q should ALWAYS be in Watts



# TODO: Copy and paste the line below into the Pycharm Terminal window to install required packages for this model. The modules will import automatically in the script. Once you install the packages, you can just run this script in Python Console.
# pip install numpy scipy pandas xlrd openpyxl

import os
import csv
import math
import xlrd
import numpy as np
import scipy
from scipy import integrate
import pandas as pd
import sys
import pickle
import openpyxl
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.drawing.image import Image
from openpyxl.drawing.colors import ColorChoice
from openpyxl.chart.marker import DataPoint
np.set_printoptions(threshold=np.inf)

# Optimization Loop TODO: Turn me on to optimize A_rsc, A_sc, and V_tank. Don't forget to indent the whole code (expect post processing) so it's enclosed in the optimization loop
# A = [3.79, 5.68, 7.57, 9.46, 11.36]
# Total_cost = np.zeros(len(A))
# for C in range(len(A)):
#     V_tank = A[C]

# Read in Data



#Read in TMY3 Data
TMY3_filename = "Data/Input Data/AK_FAIRBANKS-IAP_702610_TMY3.csv" #Locate File
T_amb = [] # Initialize empty ambient temperature list
DHI = [] # Initialize diffuse horizontal irradiance list
DNI = [] # Initialize direct normal irradiance list
vwind = [] # Initialize wind speed list
P_air = [] # Initializes air pressure list
RH = [] # Initialize relative humididty list
with open(TMY3_filename) as csvfile: # Open the CSV file
    reader = csv.reader(csvfile)

    for _ in range(19): # Skip the first 19 rows
        next(reader)
    for _ in range(8760): # Read the next 8760 rows and extract the 4th column (index 3)
        row = next(reader)

        T_amb.append(float(row[3])+273.15) #Add in TMY3 temperature data into empty list and convert from C to K
        DHI.append(float(row[12])) # Add in TMY3 Diffuse Horizontal Irradiance (W/M^2)
        DNI.append(float(row[11]))  # Add in TMY3 Direct Normal Irradiance (W/M^2)
        vwind.append(float(row[18])) # Add in TMY3 wind speed
        P_air.append(float(row[6])) # Add in TYM3 Air Pressure
        RH.append(float(row[5])) # Add in TMY3 relative humidity

# Read in heating oil price data
workbook = xlrd.open_workbook('Data/Input Data/heating_oil_prices_fairbanks.xls') # Open the workbook and select the first sheet
sheet = workbook.sheet_by_index(0)
heating_oil_price = []
for row_idx in range(1, sheet.nrows):  # First row is a header, start from second
    heating_oil_price.append(sheet.cell_value(row_idx, 1)) # Read the values from the second column into a list
heating_oil_price = np.array(heating_oil_price)
heating_oil_price = heating_oil_price.reshape(365, 24) # Shapes data into matrix

# Read in electricity cost data
workbook = xlrd.open_workbook('Data/Input Data/electricity_prices_fairbanks.xls')
sheet = workbook.sheet_by_index(0)
electricity_price = []
for row_idx in range(1, sheet.nrows):  # First row is a header, start from the second row
    electricity_price.append(sheet.cell_value(row_idx, 1)) # Read the values from the second column into a list
electricity_price = np.array(electricity_price)
electricity_price = electricity_price.reshape(365, 24) # Shapes data into matrix


# Read in Spectral Solar Irradiance Data
ext_folder = "../SMARTS_295_PC/Fairbanks OUT"  # Define the directory containing the .ext files
ext_data = []  # List of lists to store data from each .ext file
for i in range(1, 8761): # Loop through each .ext file in the specified range
    filename = os.path.join(ext_folder, f"{str(i).zfill(4)}ext.txt")
    file_data = []
    if os.path.exists(filename): # Read the file if it exists
        with open(filename, "r") as file:
            # Skip the header line if the file is not empty
            first_line = file.readline().strip()
            if first_line != "":  # Only attempt to read if there is content
                for line in file:
                    line = line.strip()
                    if line:  # Check if the line is not empty
                        parts = line.split()
                        try:
                            # Only add the second column (Global_horizn_irradiance) as a float
                            file_data.append(float(parts[1]))
                        except (ValueError, IndexError):
                            print(f"Skipping invalid line in {filename}: {line}")
            else:
                file_data = [0.0] * 1471  # Fill with zeros if the file is empty
    else:
        file_data = [0.0] * 1471  # Fill with zeros if the file is empty

    # Ensure each file_data has exactly 1471 data points by padding with zeros
    padded_data = file_data + [0.0] * (1471 - len(file_data))
    ext_data.append(padded_data)

SSI = np.array(ext_data) # Spectral solar irradiance data (List of GHI values as a function of lmbda, for each summer hour)
SSI = SSI.reshape(365, 24, 1471)
SSI = SSI * 1000 # Multiplying by 1000 to convert the irradiance from nm to um


# Read in Spectral Transmissivity Data
SAT_filename = "Data/Input Data/AM15_atm_intp.xls" #Locate File
SAT = [] # Initialize empty spectral atmospheric transmissivity list
workbook = xlrd.open_workbook(SAT_filename)
sheet = workbook.sheet_by_index(0)
for _ in range(1471): # Read the next 1471 rows and extract the first column
    SAT.append(float(sheet.cell_value(_,3))) # Add in spectral atmospheric transmissivity data for lambda = 3:.01:15 (1471 Data points)

# Read in Spectral RSC Emmisivity Data
SRSCE_filename = "Data/Input Data/RC_absorp_coeff.xls" #Locate File
SRSCE = [] # Initialize spectral RSC emmisivity list
workbook = xlrd.open_workbook(SRSCE_filename)
sheet = workbook.sheet_by_index(0)
for _ in range(1471): # Read the next 1471 rows and extract the first column
    SRSCE.append(float(sheet.cell_value(_,1))) # Add in spectral RSC emisivity data for lambda = 3:.01:15 (1471 Data points)
SRSCE = np.array(SRSCE)

# Spectral Integration axes
theta = np.linspace(0, math.pi/2, 16) # Create list of Polar Angle values from 0 to pi/2 for integration with resolution of pi/32
lmbda = np.linspace(.3, 15.01, 1471)

# Calculate the Spectral Atmospheric Emissivity
SAE = np.zeros((len(lmbda), len(theta)))
for res in range(len(lmbda)):
    for ang in range(len(theta)):
        SAE[res,ang] = 1 - SAT[res]**(1/np.cos(theta[ang])) # Creates matrix of emissivity values at varying wavelength and solar elevation angles
SAE = np.array(SAE)












# Set constants, calculate geometries, and initialize arrays

# Constants for indexing
days_in_year = 365
hours_in_day = 24
Hours_in_year = np.linspace(1, 8760, 8760) # Hour of the year
Time = np.array(pd.date_range(start='1/1/2023', end='1/1/2024', freq='h')[:-1])

# Building Material Properties and Geometry (Casey's house)

R_w = 38.976  # R-value for walls
R_r = 75.3  # R-value for roof
R_f = 50.5 # R-value for floor
U_w = .22  # U-value for windows (W/m²·K)
A_walls = 204.108  # Surface area of walls (m²)
A_roof = 277.2  # Surface area of roof (m²)
A_floor = 218.415   # Surface area of floor (m²)
A_windows = 19.6  # Total window area (m²)
V_house = 707.92 # (m3) Heating volume of the home. Use next line if not known
#V_house = ((A_walls + A_windows) * math.sqrt(A_floor))/4 # Assuming cube shaped building, calculates the rough volume of structure
Tau_windows = 0.7  # Transmittance (efficiency) of windows in transmitting solar radiation
alpha_house = 0.91  # Absorptivity of the building's surface (metal siding/roofing, thermoworks)
epsilon_house = 0.91  # Emissivity of the roof and walls

# Constants for Planck's law
C_1 = 3.742E+8
C_2 = 1.439E+4

# Constants for the rooftop solar tube heater (2021 S Yoon)
A_rsc = 5  # Optimized Area of the RSC panel in square meters
A_sc = 50  # Optimized Area of the solar collector (m^2)
F_R = 0.77  # Collector heat removal factor
tau_sc = 0.9  # Glass transmittance
alpha_sc = 0.9  # Collector absorptance
U_sc = 15.7  # Overall heat loss coefficient (W/m^2·K)

# Water Tank Variables
V_tank = 7.57 # m^3, Optimized tank size
T_tank = 271.45 # Kelvin, initial tank temperature for the first hour of the year (uses converging temperature from model)
rho_water = 997 # kg/m^3, Density of water at ambient conditions (Thermo text)
rho_glycol = 1055 # kg/m^3, Density of 50/50 (by mass) prop Glycol mix taken at -7C (averaged over design temp) (Engineering Toolbox)
cp_water = 4186 # J/kg*K, Heat capacity of water at ambient conditions
cp_glycol = 3559 # J/kg*K 50/50 mix (Engineering toolbox)
m_water = V_tank * rho_water # Kg, Mass of water in tank
m_glycol = V_tank * rho_glycol
Tfreezing_water = 273.15 # K
Tfreezing_glycol = 239.15 # Freezing temp of 50/50 PG mix

# Use this if you want to consider heat loss from an insulated tank
# Tank_radius = 1.1 # m
# Tank_height = 2 # m, Dimensions allow for a tank large enough to hold 7.57 m^3 of water
# U_tank = .3 # Heat transfer coefficient of well insulated tank
# t_insulation = .133 # m, Thickness of insulation
# Tank_wall_area = 2 * math.pi * (Tank_radius + t_insulation) * Tank_height # m^2, Calculates the surface area of the tank walls
# Tank_top_area = math.pi * (Tank_radius + t_insulation)**2 # m^2, Calculates top area of tank
# alpha_tank = .7 # Absorptivity of the water tank

# Heating Constants
LHV_oil = 44349.9 * (10**3) # J/kg, Average Lower Heating Value of No. 2 Heating Oil (Engineering Toolbox)
rho_oil = 875 # kg/m^3 Average density @ 15 degrees C (Engineering Toolbox)
gal_per_cubic_meter = 1/.003785 # (gal/m^3) Conversion Factor
boiler_efficiency = .867 # Assume constant 85% boiler efficiency
LHV_gas = 47104247.1 # J/kg, LHV of natural gas (Engineering Toolbox)
rho_gas = .777 # kg/m^3

# Emissions Constants
CO2_oil = 10.21 # KgCO2/gal
CO2_gas = .0073 # KgCO2/gal (gaseous)
CO2_electricity = .00039 # Kg/Wh

# convection coefficient
h = (3.1 + (4.1 * np.array(vwind))) # variable convection coefficient from Robert's thesis
h = h.reshape(days_in_year, hours_in_day)

# Defined Temperatures and Irradiance
T_set = 294.25 # Kelvin, initial set point temperature (70F)
T_actual = T_set # Sets the initial temp to the setpoint temperature
T_min = 292.039 # Defines minimum temp for deadband in Kelvin (66F)
T_max = 295.928 # Define maximum temp for deadband in Kelvin (73F)
T_amb = np.array(T_amb) # TMY3 Temperature Matrix
T_amb = T_amb.reshape(days_in_year, hours_in_day) # creates a matrix with 365 rows and 24 columns corresponding to hourly temperature

ASHP = 1 # Starts in January when the ASHP would be in use
WSHP = 0 # boiler and WSHP would likely be off on January 1st
boiler = 0 # These are boolean style variables which state whether a system is active or not

# Empty Heating Arrays
Q_Load = np.zeros((days_in_year,hours_in_day)) # Initialize empty heating load array
Q_Loss = np.zeros((days_in_year, hours_in_day)) # Net passive heat loss from home
Q_WSHP = np.zeros((days_in_year,hours_in_day)) # Amount of heat provided by WSHP (Wh)
Q_WSHP_Condenser = np.zeros((days_in_year,hours_in_day)) # Amount of heat pulled from tank during WSHP operation (Wh)
COP_WSHP = np.ones((days_in_year,hours_in_day)) # COP of WSHP
WSHP_expenses = np.zeros((days_in_year, hours_in_day)) # Cost to operate WSHP ($)
WSHP_electricity_consumed = np.zeros((days_in_year, hours_in_day)) # Electricity consumed by WSHP (kWh)
HC_ASHP = np.zeros((days_in_year,hours_in_day)) # Heating capacity of ASHP
HC_WSHP = np.zeros((days_in_year,hours_in_day)) # Heating capacity of WSHP

Q_boiler = np.zeros((days_in_year,hours_in_day)) # Heat provided by boiler (Wh)
boiler_expenses = np.zeros((days_in_year, hours_in_day)) # Cost to operate boiler ($)
fuel_consumed = np.zeros((days_in_year, hours_in_day)) # Amount of fuel used during operation (gal)
boiler_electricity = np.zeros((days_in_year, hours_in_day)) # Electricity consumed while running boiler (kWh)

Q_ASHP = np.zeros((days_in_year,hours_in_day)) # Heat provided by ASHP (Wh)
COP_ASHP = np.ones((days_in_year,hours_in_day)) # COP of ASHP
ASHP_expenses = np.zeros((days_in_year, hours_in_day)) # Cost to operate ASHP ($)
ASHP_electricity_consumed = np.zeros((days_in_year, hours_in_day)) # Electricity consumed by ASHP (kWh)

calculated_cost_WSHP = np.zeros((days_in_year,hours_in_day)) # What the WSHP would cost
calculated_cost_ASHP = np.zeros((days_in_year,hours_in_day)) # What the ASHP would cost
calculated_cost_boiler = np.zeros((days_in_year,hours_in_day)) # What the boiler would cost
calculated_CO2_WSHP = np.zeros((days_in_year,hours_in_day)) # WHat the system would emit if it was on
calculated_CO2_ASHP = np.zeros((days_in_year,hours_in_day))
calculated_CO2_boiler = np.zeros((days_in_year,hours_in_day))
CO2_boiler_list = np.zeros((days_in_year,hours_in_day))
CO2_ASHP_list = np.zeros((days_in_year,hours_in_day))
CO2_WSHP_list = np.zeros((days_in_year,hours_in_day))

Q_net = np.zeros((days_in_year, hours_in_day)) # Net heat gain by water tank from RSC and solar collector
Q_sc = np.zeros((days_in_year, hours_in_day)) # Heat gain from solar collector
Q_RSC = np.zeros((days_in_year, hours_in_day)) # Heat gain from RSC
Q_RSC_sun = np.zeros((days_in_year, hours_in_day))
Q_RSC_atm = np.zeros((days_in_year, hours_in_day))
Q_RSC_space = np.zeros((days_in_year, hours_in_day))
Q_RSC_conv = np.zeros((days_in_year, hours_in_day))
T_tank_list = np.zeros((days_in_year, hours_in_day)) # List to track tank temperature
T_tank_list[0,0] = T_tank # Sets the initial value
T_actual_list = np.zeros((days_in_year, hours_in_day)) # List to track inside temperature
T_actual_list[0,0] = T_actual # Sets the initial value

# Initialize lists to store each component of the home's energy balance
T_sol_air_walls_list = np.zeros((days_in_year, hours_in_day)) # Sol-air temperature (effective temperature of outside walls including radiation and convection)
T_sol_air_roof_list = np.zeros((days_in_year, hours_in_day)) # Sol-air temperature (effective temperature of outside roof including radiation and convection)
Q_walls_list = np.zeros((days_in_year, hours_in_day)) # Heat loss through walls
Q_roof_list = np.zeros((days_in_year, hours_in_day)) # Heat loss through roof
Q_floor_list = np.zeros((days_in_year, hours_in_day)) # Heat loss through floor
Q_windows_list = np.zeros((days_in_year, hours_in_day)) # Heat loss through windows
Q_ventilation_list = np.zeros((days_in_year, hours_in_day)) # Heat loss through ventilation
Q_HVAC_list = np.zeros((days_in_year, hours_in_day)) # Resultant net heating load

# Initialize solar irradiance arrays
ViewFactorHorizontal = np.zeros((days_in_year, hours_in_day))
ViewFactorVertical = np.zeros((days_in_year, hours_in_day))
ViewFactorTilted = np.zeros((days_in_year, hours_in_day))
Beta = np.zeros((days_in_year, hours_in_day))
zenith_angle = np.zeros((days_in_year, hours_in_day))
solar_altitude = np.zeros((days_in_year, hours_in_day))
DNI = np.array(DNI)
DNI = DNI.reshape(days_in_year, hours_in_day)
DHI = np.array(DHI)
DHI = DHI.reshape(days_in_year, hours_in_day)

# Initialize ventilation arrays
RH = np.array(RH) # Relative Humidity
P_air = np.array(P_air) # Air Pressure
P_air = P_air.reshape((days_in_year, hours_in_day))
RH = RH.reshape(days_in_year, hours_in_day)
omega = np.zeros((days_in_year, hours_in_day)) # Absolute humidity
Psat_vapor = np.zeros((days_in_year, hours_in_day)) # Water vapor saturation pressure
P_vapor = np.zeros((days_in_year, hours_in_day)) # Partial pressure of water vapor in air
h_dry_air = np.zeros((days_in_year, hours_in_day)) # specific enthalpy of dry air
h_vapor = np.zeros((days_in_year, hours_in_day)) # specific enthalpy of water vapor in air
h_outside_air = np.zeros((days_in_year, hours_in_day)) # specific enthalpy of moist outside air
h_inside_air = np.zeros((days_in_year, hours_in_day)) # specific enthalpy of moist inside air at varying temp
h_set_air = np.zeros((days_in_year, hours_in_day)) # specific enthalpy of moist inside air at setpoint temp
mdot_air = np.zeros((days_in_year, hours_in_day)) # mass flow rate of air
m_air = np.zeros((days_in_year, hours_in_day)) # mass of home's air
rho_air = np.zeros((days_in_year, hours_in_day)) # Moist air density
cp_moist = np.zeros((days_in_year, hours_in_day)) # Moist air specific heat capacity

# Constants for ventilation - see Thermo text and written derivation
ACH = .35 # Assumes the volume of air must be changed .35 times per hour through ventilation (ASHRAE 62.1 page 7 minimum)
vent_rate = ACH * V_house # (m^3/h) calculates the ventilation rate by volume
cp_air = 1005 # (J/kg*K)  Thermo text
cp_watervapor = 1840 # (J/kg*K) Averaged over ASHRAE design temperature range (-40F to 78F)
R_air = 286.9 # (J/kg*K) Engineering Toolbox "Moist Air"
R_vapor = 461.5 # (J/kg*K) Engineering Toolbox "Moist Air"
latent_heat = 2500.9*1000 # (J/kg) Latent heat of vaporization of water taken at 0 degrees C (does not vary much over temp ranges)
sensible_heat = cp_watervapor * (T_amb - 273.15) # (J/kg) Calculates sensible heat using celcius
P_atm = 101325  # (Pa) Atmospheric pressure at the boiling point of water
T_boiling = 373.15  # Boiling point of water in Kelvin

# Thermal mass calculations
thickness_floor = 0.15  # (m) Floor thickness (concrete slab)
rho_concrete = 2300  # (kg/m^3) Density of concrete
c_concrete = 0.653  # (kJ/kg·K) Specific heat capacity of concrete

thickness_gypsum = 0.012  # (m) Thickness of gypsum board
rho_gypsum = 800  # (kg/m^3) Density of gypsum
c_gypsum = 1.09  # (kJ/kg·K) Specific heat capacity of gypsum

thickness_glass = 0.005  # (m) Thickness of glass
rho_glass = 2230  # (kg/m^3) Density of glass
c_glass = 0.8  # (kJ/kg·K) Specific heat capacity of glass

thickness_wood = .28575 # (m) Thickness of 2x12 framing
rho_wood = 513 # (kg/m^3) Density of soft wood
c_wood = 1.38 # (kJ/kg*K) Specific heat capacity of soft wood

thickness_insulation = .28575 # (m) Thickness of wall/roof insulation
rho_insulation = 50 # (kg/m^3) Density of wood fiber batt insulation (Greenspec)
c_insulation = 2.1 # (kJ/kg*K) Specific heat of wood fiber batt insulation (Greenspec)

thickness_plywood = .0127 # (m) Thickness of plywood
rho_plywood = 545 # (kg/m^3) Density of plywood
c_plywood = 1.21 # (kJ/kg*K) Specific heat of plywood

thickness_siding = .00061 # (m) Thickness of 24 guage sheet metal cladding
rho_siding = 7851 # (kg/m^3) Density of steel sheet metal
c_siding = .49 # (kJ/kg*K) Specific heat of steel sheet metal

C_floor = (thickness_floor * rho_concrete * c_concrete * A_floor) / 3.6 # (Wh/K) Floor heat capacity (concrete slab)
C_framing = (thickness_wood * rho_wood * c_wood * (A_walls + A_roof) * .09375) / 3.6 # (Wh/K) Heat capacity of wood framing, scaling factor for approx stud spacing
C_insulation = (thickness_insulation * rho_insulation * c_insulation * (A_roof + A_walls)) / 3.6 # (Wh/K) Heat capacity of insulation
C_drywall = (thickness_gypsum * rho_gypsum * c_gypsum * (A_walls + A_floor)) / 3.6 # (Wh/K) Heat capacity of drywall on walls and ceiling
C_plywood = (thickness_plywood * rho_plywood * c_plywood * (A_walls + A_roof)) / 3.6 # (Wh/K) Heat capacity of plywood sheating
C_siding = (thickness_siding * rho_siding * c_siding * (A_walls + A_roof)) / 3.6 # (Wh/K) Heat capacity of the metal siding
C_windows = (thickness_glass * rho_glass * c_glass * A_windows) / 3.6  # (Wh/K)  Window heat capacity (glass)
C_building = (C_floor + C_framing + C_insulation + C_drywall + C_plywood + C_siding + C_windows) # (Wh/K) Building Heat capacity




# Calculate Solar Irradiance Values on vertical and horizontal surfaces

# Calculate the view factor for horizontal and vertical surfaces
for day in range(days_in_year):
    for hour in range(hours_in_day):
        latitude = 1.131 # Latitude of Fairbanks in radians
        tilt_angle = latitude
        solar_declination = 23.45*np.sin((math.pi/180)*360*((284+day)/365))*math.pi/180 # Calculates solar declination angle of day (Macgrath EQ# 2-11)
        declination_hr = np.cos(2*math.pi*((hour+1+11)/24))+(np.sin(latitude)*np.sin(solar_declination)) # Calculates solar declination angle of hour (Macgrath EQ# 2-12)
        zenith_angle[day,hour] = np.arccos(np.cos(latitude)*np.cos(declination_hr)) # Calculates the solar zenith angle of each hour (Macgrath EQ# 2-11)
        Beta[day, hour] = zenith_angle[day, hour] - tilt_angle # Angle for view factor for tilted surface (see written derivation)
        ViewFactorTilted[day, hour] = np.cos(Beta[day, hour])
        ViewFactorHorizontal[day,hour] = np.cos(zenith_angle[day,hour]) # Calculates the horizontal view factor (cosine of zenith angle) for each hour of each day and stores in matrix
        ViewFactorVertical[day,hour] = np.sin(zenith_angle[day,hour])  # Calculates the vertical view factor (sine of the zenith angle) for each hour of each day and stores in matrix

# Calculate Irradiation from TMY3 Data
HI = DNI*ViewFactorHorizontal + DHI # (W/m^2) Radiation felt by a horizontal surface
VI = DNI*ViewFactorVertical + DHI/2 # (W/m^2) Radiation felt by a vertical surface (Assumes vertical surface recieves 1/2 the diffuse component)
TI = DNI*ViewFactorTilted + DHI*((np.pi - tilt_angle)/np.pi) # Ignores Azimuthal view factor due to cylindrical tube shape



# Building Energy balance function
def energy_balance(T_amb, T_actual, VI, HI, h, mdot_air, h_outside_air, h_inside_air, day, hour): # Defines the home's heat loss in Watts
    T_sol_air_walls = T_amb + (alpha_house * .5 * VI) / h  # Simplified sol-air temperature felt by outside walls (Does not account for radiation emitted to atmosphere)
    T_sol_air_roof = T_amb + (alpha_house * HI) / h  # Simplified sol-air temperature felt by horizontal roof (Does not account for radiation emitted to atmosphere)
    Q_walls = ((T_actual - T_sol_air_walls) * A_walls) / R_w # Heat loss through walls
    Q_roof = ((T_actual - T_sol_air_roof) * A_roof) / R_r # Heat loss through roof
    Q_floor = ((T_actual - T_amb) * A_floor) / R_f # Heat loss through floor
    Q_windows = (U_w * A_windows * (T_actual - T_amb)) - (.5 * VI * A_windows * Tau_windows * .6) # Net heat loss through windows via solar gain and convective loss. Assumes windows shade 40% of the incoming solar irradiance
    Q_ventilation = mdot_air * (h_inside_air - h_outside_air) # When the outside air is colder, we heat it up to ventilate. When it's warmer, we take heat out to balance. Computes heating/cooling required for ventilation
    Q_HVAC = Q_walls + Q_roof + Q_floor + Q_windows + Q_ventilation # Returns net rate of heat loss from the home in Watts

    # Store each component in its respective list
    T_sol_air_walls_list[day, hour] = T_sol_air_walls
    T_sol_air_roof_list[day, hour] = T_sol_air_roof
    Q_walls_list[day, hour] = Q_walls
    Q_roof_list[day, hour] = Q_roof
    Q_floor_list[day, hour] = Q_floor
    Q_windows_list[day, hour] = Q_windows
    Q_ventilation_list[day, hour] = Q_ventilation
    Q_HVAC_list[day, hour] = Q_HVAC # Records the net rate of heat loss from home in Watts

    return Q_HVAC


# Building energy Balance Loop




for day in range(days_in_year):
    for hour in range(hours_in_day):

        # Goff-Gratch equation to calculate saturation pressure (Accurate over the design temperature range -40F to 78F)
        term1 = -7.90298 * ((T_boiling / T_amb[day, hour]) - 1)
        term2 = 5.02808 * math.log10(T_amb[day, hour] / T_boiling)
        term3 = -1.3816e-7 * (10 ** (11.344 * (1 - T_amb[day, hour] / T_boiling)) - 1)
        term4 = 8.1328e-3 * (10 ** (-3.49149 * ((T_boiling / T_amb[day, hour]) - 1)) - 1)
        log_psat = term1 + term2 + term3 + term4 # Logarithmic part of the equation
        Psat_vapor[day, hour] = P_atm * (10 ** log_psat) # Calculate Psat in Pascals

        # Calculate the enthalpy of the air
        P_vapor[day, hour] = Psat_vapor[day, hour] * RH[day, hour] # Vapor Pressure (Thermo text EQ #14-9)
        omega[day,hour] = (.622 * P_vapor[day, hour])/(P_air[day, hour] - P_vapor[day, hour]) # Absolute humidity (Thermo text EQ # 14-11b)
        rho_air[day, hour] =  ((P_air[day, hour] / (R_air * T_amb[day, hour]))* (1 + omega[day, hour])) / (1 + ((omega[day, hour] * R_vapor)/ R_air)) # (Kg/m^3) Calculates density of moist air (Engineering Toolbox "Moist Air")
        mdot_air[day, hour] = (vent_rate * rho_air[day, hour]) / 3600  # (kg/s) Calculates the mass flow rate of air
        m_air[day, hour] = rho_air[day, hour] * V_house # (Kg) Mass of home's moist air
        h_dry_air[day, hour] = cp_air * (T_amb[day, hour] - 273.15) # (J/kg) enthalpy of the dry air
        h_vapor[day, hour] = latent_heat + sensible_heat[day, hour] # (J/kg) Calculates enthalpy of water vapor (Thermo text EQ 14-4)
        h_outside_air[day, hour] = h_dry_air[day, hour] + (h_vapor[day, hour] * omega[day, hour]) # (J/kg) specific enthalpy of moist air (outside) (Thermo text EQ # 14-12)
        h_inside_air[day, hour] = (cp_air * (T_actual - 273.15)) + (latent_heat + (cp_watervapor * (T_actual - 273.15))) * omega[day,hour] # (J/kg) specific enthalpy of moist air at varying temperature (inside) (Thermo text EQ # 14-12)
        h_set_air[day, hour] = (cp_air * (T_set - 273.15)) + (latent_heat + (cp_watervapor * (T_set - 273.15))) * omega[day,hour] # (J/kg) specific enthalpy of moist air at setpoint temperature (inside) (Thermo text EQ # 14-12)
        cp_moist[day, hour] = cp_air + (omega[day, hour] * cp_watervapor) # (J/kg*K) Defines specific heat capacity of the moist air

        # Calculate the home heat loss, perform home energy balance

        Q_Loss[day, hour] = energy_balance(T_amb[day, hour], T_actual, VI[day, hour], HI[day, hour], h[day, hour], mdot_air[day, hour], h_outside_air[day, hour], h_inside_air[day, hour], day, hour)  # Net heat loss from home in Watts

        if (T_actual < T_min) or (T_actual > T_max):  # Tests to see if actual temperature is outside of deadband
            delta_H_building = C_building * (T_set - T_actual)  # (Wh) Change in Enthalpy of the building
            delta_H_air = m_air[day, hour] * (h_set_air[day, hour] - h_inside_air[day, hour])/3600  # (Wh) Change in Enthalpy of the inside air (divides by 3600 to convert from J to Wh)
            LOAD = delta_H_air + delta_H_building + Q_Loss[day, hour]  # (Wh) Gives the HVAC heating load to return the building to it's setpoint temperature. Here, Q_loss is in Watt hours, which is numerically the same as it's value in Watts



            # BELOW 20kW
            if (abs(LOAD) <= 20000):
                Q_Load[day, hour] += LOAD



            # BELOW 40kW
            elif (abs(LOAD) <= 40000):
                if LOAD > 0:
                    Q_Load[day, hour] += 20000
                elif LOAD < 0:
                    Q_Load[day, hour] += -20000

                # Same day, next hour
                if hour < 23:
                    next_hour = hour + 1
                    if LOAD > 0:
                        Q_Load[day, next_hour] += LOAD - 20000
                    elif LOAD < 0:
                        Q_Load[day, next_hour] += LOAD + 20000

                # Need to iterate to next day
                elif (hour == 23) and (day < 364):
                    next_day = day + 1
                    next_hour = 0
                    if LOAD > 0:
                        Q_Load[next_day, next_hour] += LOAD - 20000
                    elif LOAD < 0:
                        Q_Load[next_day, next_hour] += LOAD + 20000

                # Need to iterate to next year
                elif (hour == 23) and (day == 364):
                    next_day = 0
                    next_hour = 0
                    if LOAD > 0:
                        Q_Load[next_day, next_hour] += LOAD - 20000
                    elif LOAD < 0:
                        Q_Load[next_day, next_hour] += LOAD + 20000







            # BELOW 60kW
            elif (abs(LOAD) <= 60000):
                if LOAD > 0:
                    Q_Load[day, hour] += 20000
                elif LOAD < 0:
                    Q_Load[day, hour] += -20000

                # NEXT HOUR

                # Same day, next hour
                if hour < 23:
                    next_hour = hour + 1
                    if LOAD > 0:
                        Q_Load[day, next_hour] += 20000
                    elif LOAD < 0:
                        Q_Load[day, next_hour] += -20000

                # Need to iterate to next day
                elif (hour == 23) and (day < 364):
                    next_day = day + 1
                    next_hour = 0
                    if LOAD > 0:
                        Q_Load[next_day, next_hour] += 20000
                    elif LOAD < 0:
                        Q_Load[next_day, next_hour] += -20000

                # Need to iterate to next year
                elif (hour == 23) and (day == 364):
                    next_day = 0
                    next_hour = 0
                    if LOAD > 0:
                        Q_Load[next_day, next_hour] += 20000
                    elif LOAD < 0:
                        Q_Load[next_day, next_hour] += -20000

                # THIRD HOUR

                # Same day, third hour
                if hour < 22:
                    third_hour = hour + 2

                    if LOAD > 0:
                        Q_Load[day, third_hour] += LOAD - 40000
                    elif LOAD < 0:
                        Q_Load[day, third_hour] += LOAD + 40000

                # Need to iterate to next day
                elif (hour >= 22) and (day < 364):
                    next_day = day + 1
                    if hour == 22:
                        third_hour = 0
                    elif hour == 23:
                        third_hour = 1

                    if LOAD > 0:
                        Q_Load[next_day, third_hour] += LOAD - 40000
                    elif LOAD < 0:
                        Q_Load[next_day, third_hour] += LOAD + 40000

                # Need to iterate to next year
                elif (hour >= 22) and (day == 364):
                    next_day = 0
                    if hour == 22:
                        third_hour = 0
                    elif hour == 23:
                        third_hour = 1

                    if LOAD > 0:
                        Q_Load[next_day, third_hour] += LOAD - 40000
                    elif LOAD < 0:
                        Q_Load[next_day, third_hour] += LOAD + 40000






            # ABOVE 60kW
            else:
                if LOAD > 0:
                    Q_Load[day, hour] += 20000
                elif LOAD < 0:
                    Q_Load[day, hour] += -20000

                # NEXT HOUR

                # Same day, next hour
                if hour < 23:
                    next_hour = hour + 1
                    if LOAD > 0:
                        Q_Load[day, next_hour] += 20000
                    elif LOAD < 0:
                        Q_Load[day, next_hour] += -20000

                # Need to iterate to next day
                elif (hour == 23) and (day < 364):
                    next_day = day + 1
                    next_hour = 0
                    if LOAD > 0:
                        Q_Load[next_day, next_hour] += 20000
                    elif LOAD < 0:
                        Q_Load[next_day, next_hour] += -20000

                # Need to iterate to next year
                elif (hour == 23) and (day == 364):
                    next_day = 0
                    next_hour = 0
                    if LOAD > 0:
                        Q_Load[next_day, next_hour] += 20000
                    elif LOAD < 0:
                        Q_Load[next_day, next_hour] += -20000

                # THIRD HOUR

                # Same day, third hour
                if hour < 22:
                    third_hour = hour + 2
                    if LOAD > 0:
                        Q_Load[day, third_hour] += 20000
                    elif LOAD < 0:
                        Q_Load[day, third_hour] += -20000

                # Need to iterate to next day
                elif (hour >= 22) and (day < 364):
                    next_day = day + 1
                    if hour == 22:
                        third_hour = 0
                    elif hour == 23:
                        third_hour = 1

                    if LOAD > 0:
                        Q_Load[next_day, third_hour] += 20000
                    elif LOAD < 0:
                        Q_Load[next_day, third_hour] += -20000

                # Need to iterate to next year
                elif (hour >= 22) and (day == 364):
                    next_day = 0

                    if hour == 22:
                        third_hour = 0

                    elif hour == 23:
                        third_hour = 1

                    if LOAD > 0:
                        Q_Load[next_day, third_hour] += 20000
                    elif LOAD < 0:
                        Q_Load[next_day, third_hour] += -20000

                # FOURTH HOUR

                # Same day, fourth hour
                if hour < 21:
                    fourth_hour = hour + 3
                    if LOAD > 0:
                        Q_Load[day, fourth_hour] += LOAD - 60000
                    elif LOAD < 0:
                        Q_Load[day, fourth_hour] += LOAD + 60000

                # Need to iterate to next day
                elif hour >= 21 and day < 364:
                    next_day = day + 1
                    if hour == 21:
                        fourth_hour = 0
                    elif hour == 22:
                        fourth_hour = 1
                    elif hour == 23:
                        fourth_hour = 2

                    if LOAD > 0:
                        Q_Load[next_day, fourth_hour] += LOAD - 60000
                    elif LOAD < 0:
                        Q_Load[next_day, fourth_hour] += LOAD + 60000

                # Need to iterate to next year
                elif (hour >= 21) and (day == 364):
                    next_day = 0
                    if hour == 21:
                        fourth_hour = 0
                    elif hour == 22:
                        fourth_hour = 1
                    elif hour == 23:
                        fourth_hour = 2

                    if LOAD > 0:
                        Q_Load[next_day, fourth_hour] += LOAD - 60000
                    elif LOAD < 0:
                        Q_Load[next_day, fourth_hour] += LOAD + 60000



        # Balance for inside air temperature
        T_actual_list[day, hour] = (T_actual - 273.15) * (9/5) + 32  # (F) Records the temperature of that iteration before recalculating for next iteration
        T_actual = T_actual + (Q_Load[day, hour] - Q_Loss[day, hour]) / ((m_air[day, hour] * cp_moist[day, hour])/3600 + C_building) # Computes the new inside air temperature. Divides by 3600 to convert from J to Wh










# Energy Balance functions for the RSC and the Solar Collector

# Solar radiation component
def calc_solar_radiation(SSI): # Defines Absorbed Heat flux from sun
    integrand = SRSCE * SSI # Note that SSI here is Global Horizontal Irradiance, so view factor is accounted for
    return integrate.trapezoid(integrand, lmbda) # Uses trapezoid rule to integrate over wavelength range (Magrath EQ# 2-13)

# Convection component
def calc_convection(h, T_amb, T_tank): # Defines Heat flux entering RSC panel from convection
    return h * (T_amb - T_tank) # (Magrath EQ# 2-14)

# Atmospheric radiation component
def calc_atmospheric_radiation(T_amb): # Defines Heat flux absorbed from atmoshperic radiation
    B = (C_1 * lmbda ** -5) / (np.exp(C_2 / (lmbda * T_amb)) - 1)  # calculates blackbody spectral emissive power by Planck's law
    integrand = np.zeros((len(lmbda), len(theta)))
    for res in range(len(lmbda)):
        for ang in range(len(theta)):
            integrand[res, ang] = SRSCE[res] * SAE[res, ang] * B[res] * np.sin(2 * theta[ang]) # Creates 2D grid of data for integration
    integrate_over_lmbda = integrate.trapezoid(integrand, lmbda, axis=0) # integrates over lambda from .3 to 15um
    integrate_over_theta = integrate.trapezoid(integrate_over_lmbda, theta) # integrates over theta (Polar angle) from 0 to pi/2
    return integrate_over_theta # Uses nested trapezoid rule to perform double integral over Polar angle and EM spectrum (H&MT Page 780)

# Radiation to deep space component
def calc_radiation_to_space(T_tank): # Heat flux loss to space
    B = (C_1 * lmbda ** -5) / (np.exp(C_2 / (lmbda * T_tank)) - 1)  # Calculates emissive power by Planck's law
    integrand = SRSCE * B
    return integrate.trapezoid(integrand, lmbda) # Uses trapezoid rule to integrate across wavelength range (H&MT Page 794)

# Net RSC heat gain calculation
def RSC_net(h, T_tank, T_amb, SSI):
    Q_sun = calc_solar_radiation(SSI)
    Q_conv = calc_convection(h, T_amb, T_tank)
    Q_atm = calc_atmospheric_radiation(T_amb)
    Q_space = calc_radiation_to_space(T_tank)
    Q_RSC = (Q_sun + Q_conv + Q_atm - Q_space) * A_rsc

    Q_RSC_sun[day, hour] = Q_sun * A_rsc
    Q_RSC_atm[day, hour] = Q_atm * A_rsc
    Q_RSC_space[day, hour] = Q_space * A_rsc
    Q_RSC_conv[day, hour] = Q_conv * A_rsc

    return Q_RSC

# Solar Tube Heater
def Solar_Collector(T_tank, T_amb, TI): # Calculate the useful energy gain of the solar thermal collector
    Q_sc = A_sc * (F_R * tau_sc * alpha_sc * TI - F_R * U_sc * (T_tank - T_amb))
    return Q_sc
















# Tank Energy Balance
# MAIN LOOP

for day in range(days_in_year):
    for hour in range(hours_in_day):

        # Call RSC and SC Functions
        Q_sc[day, hour] = Solar_Collector(T_tank, T_amb[day, hour], TI[day, hour])  # Calls the solar collector energy balance function
        Q_RSC[day, hour] = RSC_net(h[day, hour], T_tank, T_amb[day, hour], SSI[day, hour]) # Calls the RSC energy balance function

        # Keep each system from wasting energy to environment
        if Q_sc[day, hour] < 0:
            Q_sc[day, hour] = 0  # If Solar Collector loses energy to the environment, it is immediately shut off for that hour

        if Q_RSC[day, hour] > 0:
            Q_RSC[day, hour] = 0  # If RSC gains energy from environment, it is immediately shut off for that hour

        # Calculate the COP of ASHP(T_amb) and WSHP(T_tank). Pulls directly from product data
        # Note that these equations assume a constant inlet temperature to the ASHP, neglecting any variance. COP is only a functon of ambient/tank temperature.

        # Heating Mode

        if Q_Load[day, hour] >= 0:
            if (T_amb[day, hour] - 273.15) >= -30 and (T_amb[day, hour] - 273.15) <= 38:
                COP_ASHP[day, hour] = abs((.0004 * ((T_amb[day, hour] - 273.15)**2)) + (.0637 * (T_amb[day, hour] - 273.15)) + 2.8325) # Valid for ASHP heating from -30C to 38C
                HC_ASHP[day, hour] = 1000 * abs((.2463 * (T_amb[day, hour] - 273.15)) + 13.363)

            elif (T_amb[day, hour] - 273.15) < -30:
                COP_ASHP[day, hour] = 1 # If colder than -30C assume COP becomes 1
                HC_ASHP[day, hour] = 0 # Inoperable lower than -30C

            elif (T_amb[day, hour] - 273.15) > 38:
                COP_ASHP[day, hour] = abs((.0004 * (38**2)) + (.0637 * 38) + 2.8325) # If T_amb > 38C, use COP @ 38C
                HC_ASHP[day, hour] = 1000 * abs((.2463 * 38) + 13.363)

            if (T_tank - 273.15) >= -1.1 and (T_tank - 273.15) <= 26.7:
                COP_WSHP[day, hour] = abs(((-9*10**-6) * ((T_tank - 273.15)**3)) + (.0001 * ((T_tank - 273.15)**2)) + (.0562 * (T_tank - 273.15)) + 3.4538) # Valid for WSHP heating from -1.1C to 26.7C
                HC_WSHP[day, hour] = 1000 * abs(((-2*10**-5) * ((T_tank - 273.15)**3)) + (.003 * ((T_tank - 273.15)**2)) + (.3413 * (T_tank - 273.15)) + 14.072)

            elif (T_tank - 273.15) < -1.1:
                COP_WSHP[day, hour] = 1  # If T_tank < -1.1, use COP @ lowest operational temp
                HC_WSHP[day, hour] = 0    # If T_tank < -1.1, use COP @ lowest operational temp

            elif (T_tank - 273.15) > 26.7:
                COP_WSHP[day, hour] = abs(((-9 * 10 ** -6) * (26.7 ** 3)) + (.0001 * (26.7 ** 2)) + (.0562 * 26.7) + 3.4538) # If T_tank > 26.7C, use COP @ 26.7C
                HC_WSHP[day, hour] = 1000 * abs(((-2 * 10 ** -5) * (26.7 ** 3)) + (.003 * (26.7 ** 2)) + (.3413 * 26.7) + 14.072)

        # Cooling Mode - No need to consider cooling capacity

        elif Q_Load[day, hour] < 0:
            if (T_amb[day, hour] - 273.15) >= 28 and (T_amb[day, hour] - 273.15) <= 35:
                COP_ASHP[day, hour] = abs((-.04 * (T_amb[day, hour] - 273.15) + 4)) # Valid for ASHP heating from 28C to 35C

            elif (T_amb[day, hour] - 273.15) < 28:
                COP_ASHP[day, hour] = abs((-.04 * 28 + 4)) # If its colder than 28C take COP at 28C

            elif (T_amb[day, hour] - 273.15) > 35:
                COP_ASHP[day, hour] = 1 # If its warmer than 35 assume COP becomes 1

            if (T_tank - 273.15) >= 10 and (T_tank - 273.15) <= 43.3:
                COP_WSHP[day, hour] = abs(((-1*10**-5) * ((T_tank - 273.15)**3)) + (.0025 * ((T_tank - 273.15)**2)) - (.2173 * (T_tank - 273.15)) + 8.5919) # Curve fit

            elif (T_tank - 273.15) < 10:
                COP_WSHP[day, hour] = abs(((-1*10**-5) * (10**3)) + (.0025 * (10**2)) - (.2173 * 10) + 8.5919) # If it's colder than 10C, take the COP at 10C

            elif (T_tank - 273.15) > 43.3:
                COP_WSHP[day, hour] = 1 # Assume that when the temp is warm, AC COP becomes 1






       # Optimize HVAC system based on cost
        if Q_Load[day, hour] != 0:
            min_cost = 100 # ($) Initialize minimum cost variable
            min_emission = 100 # (Kg) Initialize minimum emissions variable
            cost_WSHP = (((abs(Q_Load[day, hour]) / COP_WSHP[day, hour]) * (1/1000)) + .2) * electricity_price[day, hour] # Cost to heat/cool with WSHP for that hour. Includes 200W to run pumps/valves
            cost_ASHP = (((abs(Q_Load[day, hour]) / COP_ASHP[day, hour]) * (1/1000)) + .2) * electricity_price[day, hour]  # Cost to heat/cool with ASHP for that hour. Includes 200W to run pumps/valves (Arctic Heat Pump EVI, pg 7)
            cost_boiler = ((heating_oil_price[day, hour] * gal_per_cubic_meter * (1 / rho_oil) * (abs(Q_Load[day, hour]) * 3600)) / (LHV_oil * boiler_efficiency)) + (.2 * electricity_price[day, hour])  # Cost to heat/cool with boiler for that hour (includes assumed 200 Watts of electricity to run motor/pump)
            CO2_WSHP = CO2_electricity * (cost_WSHP / electricity_price[day, hour]) * 1000 # (Kg) Calculates Co2 emmitted that hour
            CO2_ASHP = CO2_electricity * (cost_ASHP / electricity_price[day, hour]) * 1000 # (Kg)
            CO2_boiler = CO2_oil * ((cost_boiler - (.2 * electricity_price[day, hour])) / heating_oil_price[day, hour]) + (200 * CO2_electricity) # (Kg) Includes emissions resulting from 200W auxillary power

            # Record Calculated Costs for comparison
            calculated_cost_WSHP[day, hour] = cost_WSHP
            calculated_cost_ASHP[day, hour] = cost_ASHP
            calculated_cost_boiler[day, hour] = cost_boiler
            calculated_CO2_WSHP[day, hour] = CO2_WSHP
            calculated_CO2_ASHP[day, hour] = CO2_ASHP
            calculated_CO2_boiler[day, hour] = CO2_boiler

# TODO: This Block for Cost Optimization. Turn off for emissions optimization

            # WSHP LOGIC
            if day >= 120 and day <= 240: # During summer heating season
                if (cost_WSHP <= min_cost):
                    min_cost = cost_WSHP  # WSHP becomes the lowest cost to beat
                    WSHP = 1  # WSHP will run for this hour
                    ASHP = 0
                    boiler = 0
                else:
                    WSHP = 0
            else: # This outer if-else branch allows WSHP to be used for cooling when water is extremely cold during the summer
                if (cost_WSHP <= min_cost and T_tank > (-1.1 + 273.15)): # Ensure WSHP is operable under tank temperature and is the cheapest system to run
                    min_cost = cost_WSHP # WSHP becomes the lowest cost to beat
                    WSHP = 1 # WSHP will run for this hour
                    ASHP = 0
                    boiler = 0
                else:
                    WSHP = 0

             # ASHP LOGIC
            if (cost_ASHP < min_cost and T_amb[day, hour] >= (-30 + 273.15)): # Shuts off ASHP when colder than -30C
                min_cost = cost_ASHP # ASHP becomes the lowest cost to beat
                ASHP = 1 # ASHP will run for this hour
                boiler = 0
                WSHP = 0
            else:
                ASHP = 0

            # BOILER LOGIC
            if (cost_boiler < min_cost and Q_Load[day, hour] > 0): # Make sure there is a heating load and the boiler is the cheapest system to run
                min_cost = cost_boiler # boiler cost becomes the lowest cost
                boiler = 1 # boiler will run for this hour
                ASHP = 0
                WSHP = 0
            else:
                boiler = 0

            if (WSHP == 0 and ASHP == 0 and Q_Load[day, hour] > 0): # Backup heating plan if neither ASHP or WSHP are viable
                boiler = 1


        else:  # Backup check to turn off all systems if there is no heating load
            WSHP = 0
            ASHP = 0
            boiler = 0

# TODO: This Block for Emissions Optimization. Turn off for Cost optimization.

            # WSHP LOGIC
        #     if day >= 120 and day <= 240:  # During summer heating season
        #         if (CO2_WSHP <= min_emission):
        #             min_emission = CO2_WSHP  # WSHP becomes the lowest emission to beat
        #             WSHP = 1  # WSHP will run for this hour
        #             ASHP = 0
        #             boiler = 0
        #         else:
        #             WSHP = 0
        #     else:  # This outer if-else branch allows WSHP to be used for cooling when water is extremely cold during the summer
        #         if (CO2_WSHP <= min_emission and T_tank > (-1.1 + 273.15)):  # Ensure WSHP is operable under tank temperature and is the cheapest system to run
        #             min_emission = CO2_WSHP  # WSHP becomes the lowest cost to beat
        #             WSHP = 1  # WSHP will run for this hour
        #             ASHP = 0
        #             boiler = 0
        #         else:
        #             WSHP = 0
        #
        #     # ASHP LOGIC
        #     if (CO2_ASHP < min_emission and T_amb[day, hour] >= (-30 + 273.15)):  # Shuts off ASHP when colder than -35C
        #         min_emission = CO2_ASHP  # ASHP becomes the lowest cost to beat
        #         ASHP = 1  # ASHP will run for this hour
        #         boiler = 0
        #         WSHP = 0
        #     else:
        #         ASHP = 0
        #
        #     # BOILER LOGIC
        #     if (CO2_boiler < min_emission and Q_Load[day, hour] > 0):  # Make sure there is a heating load and the boiler is the cheapest system to run
        #         min_emission = CO2_boiler  # boiler cost becomes the lowest cost
        #         boiler = 1  # boiler will run for this hour
        #         ASHP = 0
        #         WSHP = 0
        #     else:
        #         boiler = 0
        #
        #     if (WSHP == 0 and ASHP == 0 and Q_Load[day, hour] > 0):  # Backup heating plan if neither ASHP or WSHP are viable
        #         boiler = 1
        #
        # else:  # Backup check to turn off all systems if there is no heating load
        #     WSHP = 0
        #     ASHP = 0
        #     boiler = 0



        # Assign the heating load to the appropriate HVAC system

        if WSHP == 1: # Heat is only pulled from tank when WSHP is active
            if HC_WSHP[day, hour] >= Q_Load[day, hour]: # If the HP can provide the heating load, let it
                Q_WSHP[day, hour] = Q_Load[day, hour] # Records the heat supplied by WSHP
                WSHP_expenses[day, hour] = cost_WSHP # Records the cost of running the WSHP for that hour
                WSHP_electricity_consumed[day, hour] = cost_WSHP / electricity_price[day, hour] # (kWh) Records the electricity consumed by WSHP
                CO2_WSHP_list[day, hour] = CO2_WSHP # (Kg) Records the CO2 associated with that hour
                Q_WSHP_Condenser[day, hour] = Q_Load[day, hour] - (Q_Load[day, hour] / COP_WSHP[day, hour]) # Calculates the amount of heat transferred into WSHP's condenser. Subtracts the 200 Watts of auxiliary power before calculating.
            else: # If the WSHP is the best but can't meet the heating load, give the rest to the boiler
                Q_WSHP[day, hour] = HC_WSHP[day, hour]
                cost_WSHP = (((abs(Q_WSHP[day, hour]) / COP_WSHP[day, hour]) * (1 / 1000)) + .2) * electricity_price[day, hour]
                WSHP_expenses[day, hour] = cost_WSHP
                WSHP_electricity_consumed[day, hour] = cost_WSHP / electricity_price[day, hour]
                CO2_WSHP = CO2_electricity * (cost_WSHP / electricity_price[day, hour]) * 1000
                CO2_WSHP_list[day, hour] = CO2_WSHP
                Q_WSHP_Condenser[day, hour] = Q_WSHP[day, hour] - (Q_WSHP[day, hour] / COP_WSHP[day, hour])

                Q_boiler[day, hour] = Q_Load[day, hour] - Q_WSHP[day, hour]
                cost_boiler = ((heating_oil_price[day, hour] * gal_per_cubic_meter * (1 / rho_oil) * (abs(Q_boiler[day, hour]) * 3600)) / (LHV_oil * boiler_efficiency)) + (.2 * electricity_price[day, hour])
                boiler_expenses[day, hour] = cost_boiler
                fuel_consumed[day, hour] = (cost_boiler - (.2 * electricity_price[day, hour])) / heating_oil_price[day, hour]
                boiler_electricity[day, hour] = (.2 * electricity_price[day, hour])
                CO2_boiler = CO2_oil * ((cost_boiler - (.2 * electricity_price[day, hour])) / heating_oil_price[day, hour]) + (200 * CO2_electricity)  # (Kg) Includes emissions resulting from 200W auxillary power
                CO2_boiler_list[day, hour] = CO2_boiler


        if ASHP == 1: # Records when ASHP is used
            if HC_ASHP[day, hour] >= Q_Load[day, hour]: # If the HP can provide the heating load, let it
                Q_ASHP[day, hour] = Q_Load[day, hour] # Records heat supplied by ASHP
                ASHP_expenses[day, hour] = cost_ASHP # Records the cost of running the ASHP for that hour
                ASHP_electricity_consumed[day, hour] = cost_ASHP / electricity_price[day, hour] # (kWh) Records the electricity consumed by ASHP
                CO2_ASHP_list[day, hour] = CO2_ASHP # (Kg) Records the CO2 associated with that hour
            else:
                Q_ASHP[day, hour] = HC_ASHP[day, hour]
                cost_ASHP = (((abs(Q_ASHP[day, hour]) / COP_ASHP[day, hour]) * (1 / 1000)) + .2) * electricity_price[day, hour]
                ASHP_expenses[day, hour] = cost_ASHP
                ASHP_electricity_consumed[day, hour] = cost_ASHP / electricity_price[day, hour]
                CO2_ASHP = CO2_electricity * (cost_ASHP / electricity_price[day, hour]) * 1000
                CO2_ASHP_list[day, hour] = CO2_ASHP

                Q_boiler[day, hour] = Q_Load[day, hour] - Q_ASHP[day, hour]
                cost_boiler = ((heating_oil_price[day, hour] * gal_per_cubic_meter * (1 / rho_oil) * (abs(Q_boiler[day, hour]) * 3600)) / (LHV_oil * boiler_efficiency)) + (.2 * electricity_price[day, hour])
                boiler_expenses[day, hour] = cost_boiler
                fuel_consumed[day, hour] = (cost_boiler - (.2 * electricity_price[day, hour])) / heating_oil_price[day, hour]
                boiler_electricity[day, hour] = (.2 * electricity_price[day, hour])
                CO2_boiler = CO2_oil * ((cost_boiler - (.2 * electricity_price[day, hour])) / heating_oil_price[day, hour]) + (200 * CO2_electricity)  # (Kg) Includes emissions resulting from 200W auxillary power
                CO2_boiler_list[day, hour] = CO2_boiler



        if boiler == 1: # Records when boiler is used
            Q_boiler[day, hour] = Q_Load[day, hour] # Records the heat supplied by boiler
            boiler_expenses[day, hour] = cost_boiler # Records the cost of running the boiler for that hour
            fuel_consumed[day, hour] = (cost_boiler - (.2 * electricity_price[day, hour])) / heating_oil_price[day, hour] # (gal) Records the heating oil consumed in gallons by the boiler. Subtracts out the cost of auxiliary power before calculating.
            boiler_electricity[day, hour] = (.2 * electricity_price[day, hour]) # Records the 200 Watts of electricity used to run motor/pump
            CO2_boiler = CO2_oil * ((cost_boiler - (.2 * electricity_price[day, hour])) / heating_oil_price[day, hour]) + (200 * CO2_electricity)
            CO2_boiler_list[day, hour] = CO2_boiler # (Kg) Records the CO2 associated with that hour

        T_tank_list[day, hour] = T_tank  # Records the value of T_tank each iteration

        # Tank Energy Balance - Note that this assumes the tank is well insulated and there is negligible heat transfer to surroundings. Assumes that heat added through mechanical work (pumps/valves) washes out with the pipe's heat loss to the environment.
        if day >= 120 and day <= 240: # During summer heating season
            Q_net[day, hour] = Q_RSC[day, hour] - Q_WSHP_Condenser[day, hour]  # Records the net energy gain/loss of the water tank. TURNS OFF Q_sc for summer cooling season May 1st - August 29th

        else:
            Q_net[day, hour] = Q_sc[day, hour] - Q_WSHP_Condenser[day, hour]  # Records the net energy gain/loss of the water tank. TURNS OFF Q_RSC for winter heating season
        T_tank = (Q_net[day, hour] * 3600) / (m_glycol * cp_glycol) + T_tank # Resets T_tank variable by calculating new tank temperature from old tank temp and energy balance



# TODO: Turn me on to Optimize RSC, SC, and Tank size. Dont forget to indent the entire code so it is enclosed in the optimization loop

# boiler_cost = 0
# ASHP_cost = 0
# WSHP_cost = 0
# for day in range(days_in_year):
#     for hour in range(hours_in_day):
#         boiler_cost += boiler_expenses[day, hour]
#         WSHP_cost += WSHP_expenses[day, hour]
#         ASHP_cost += ASHP_expenses[day, hour]
# Total_cost[C] = boiler_cost + WSHP_cost + ASHP_cost
# C = C+1






# POST PROCESSING

# Create Pandas Dataframes
df1 = pd.DataFrame({
    'Date and Time': Time,
    'Outside Air Temp (K)': T_amb.reshape(-1),
    'Tank Temperature (K)': T_tank_list.reshape(-1),
    'Inside Air Temperature (F)': T_actual_list.reshape(-1),
    'Heating Load (Wh)': Q_Load.reshape(-1),
    'WSHP Load (Wh)': Q_WSHP.reshape(-1),
    'ASHP Load (Wh)': Q_ASHP.reshape(-1),
    'Boiler Load (Wh)': Q_boiler.reshape(-1),
    'Hour of the Year (h)': Hours_in_year})

df2 = pd.DataFrame({
    'Date and Time': Time,
    'Outside Air Temp (K)': T_amb.reshape(-1),
    'Tank Temperature (K)': T_tank_list.reshape(-1),
    'WSHP cost ($)': WSHP_expenses.reshape(-1),
    'ASHP cost ($)': ASHP_expenses.reshape(-1),
    'Boiler Cost ($)': boiler_expenses.reshape(-1),
    'Pseudo WSHP Cost ($)': calculated_cost_WSHP.reshape(-1),
    'Pseudo ASHP cost ($)': calculated_cost_ASHP.reshape(-1),
    'Pseudo Boiler cost ($)': calculated_cost_boiler.reshape(-1)})

df3 = pd.DataFrame({
    'Date and Time': Time,
    'Outside Air Temp (K)': T_amb.reshape(-1),
    'Tank Temperature (K)': T_tank_list.reshape(-1),
    'WSHP electricity usage (kWh)': WSHP_electricity_consumed.reshape(-1),
    'ASHP electricity usage (kWh)': ASHP_electricity_consumed.reshape(-1),
    'Boiler fuel Usage (gal)': fuel_consumed.reshape(-1)})

df4 = pd.DataFrame({
    'Date and Time': Time,
    'Outside Air Temp (K)': T_amb.reshape(-1),
    'Tank Temperature (K)': T_tank_list.reshape(-1),
    'WSHP CO2 (Kg)': CO2_WSHP_list.reshape(-1),
    'ASHP CO2 (Kg)': CO2_ASHP_list.reshape(-1),
    'Boiler CO2 (Kg)': CO2_boiler_list.reshape(-1),
    'Pseudo WSHP CO2 (Kg)': calculated_CO2_WSHP.reshape(-1),
    'Pseudo ASHP CO2 (Kg)': calculated_CO2_ASHP.reshape(-1),
    'Pseudo Boiler CO2 (Kg)': calculated_CO2_boiler.reshape(-1)})

df5 = pd.DataFrame({
    'Date and Time': Time,
    'Outside Air Temp (K)': T_amb.reshape(-1),
    'Tank Temperature (K)': T_tank_list.reshape(-1),
    'WSHP COP': COP_WSHP.reshape(-1),
    'ASHP COP': COP_ASHP.reshape(-1)})

df6 = pd.DataFrame({
    'Date and Time': Time,
    'Outside Air Temp (K)': T_amb.reshape(-1),
    'Tank Temp (K)': T_tank_list.reshape(-1),
    'Heating Load (Wh)': Q_Load.reshape(-1),
    'Net Heat Loss (Wh)': Q_Loss.reshape(-1),
    'Heat loss through walls (Wh)': Q_walls_list.reshape(-1),
    'Heat loss through roof (Wh)': Q_roof_list.reshape(-1),
    'Heat loss through floor (Wh)': Q_floor_list.reshape(-1),
    'Heat loss through windows (Wh)': Q_windows_list.reshape(-1),
    'Heat loss through ventilation (Wh)': Q_ventilation_list.reshape(-1)})

df7 = pd.DataFrame({
    'Date and Time': Time,
    'Outside Air Temp (K)': T_amb.reshape(-1),
    'Tank Temperature (K)': T_tank_list.reshape(-1),
    'Net Heat gain to Tank (Wh)': Q_net.reshape(-1),
    'Heat gain from solar collector (Wh)': Q_sc.reshape(-1),
    'Heat gain from RSC (Wh)': Q_RSC.reshape(-1),
    'Heat loss to WSHP (Wh)': Q_WSHP_Condenser.reshape(-1)})

df8 = pd.DataFrame({
    'Date and Time': Time,
    'Outside Air Temp (K)': T_amb.reshape(-1),
    'Tank Temperature (K)': T_tank_list.reshape(-1),
    'RSC Net Heat gain (Wh)': Q_RSC.reshape(-1),
    'Convection absorbed (Wh)': Q_RSC_conv.reshape(-1),
    'Solar irradiation absorbed (Wh)': Q_RSC_sun.reshape(-1),
    'Space radiation emmitted (Wh)': Q_RSC_space.reshape(-1),
    'Atmospheric radiation absorbed (Wh)': Q_RSC_atm.reshape(-1),})


# Write data to Excel using pandas
with pd.ExcelWriter('Data/Output Data/ModelHomeData.xlsx', engine='openpyxl') as writer:
    df1.to_excel(writer, sheet_name='Loads', index=False)
    df2.to_excel(writer, sheet_name='Cost', index=False)
    df3.to_excel(writer, sheet_name='Consumption', index=False)
    df4.to_excel(writer, sheet_name='CO2 Emissions', index=False)
    df5.to_excel(writer, sheet_name='COP', index=False)
    df6.to_excel(writer, sheet_name='Home Energy Balance', index=False)
    df7.to_excel(writer, sheet_name='Tank Energy Balance', index=False)
    df8.to_excel(writer, sheet_name='RSC Energy Balance', index=False)

# Reopen the workbook to adjust columns and freeze panes
workbook = load_workbook('Data/Output Data/ModelHomeData.xlsx')

# Function to auto-adjust column width
def auto_adjust_column_width(sheet):
    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column_letter # Get the column name
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

# Apply the auto-adjust and freeze panes for each sheet
for sheet_name in ['Loads', 'Cost', 'Consumption', 'COP', 'Home Energy Balance', 'Tank Energy Balance', 'RSC Energy Balance', 'CO2 Emissions']:
    sheet = workbook[sheet_name]
    auto_adjust_column_width(sheet)
    sheet.freeze_panes = 'A2'  # Freeze the header row

# Save the workbook with the changes
workbook.save('Data/Output Data/ModelHomeData.xlsx')








# Generate Graph

# Create Scatter Plot and add it to a new sheet
def create_scatter_chart(workbook):
    from openpyxl.chart.marker import DataPoint

    # Select the worksheet where you want to create the chart
    ws = workbook.create_sheet(title='System Usage')

    # Create a ScatterChart object
    chart = ScatterChart()
    chart.title = "HVAC System Usage"
    chart.style = 13
    chart.x_axis.title = "Hour of the Year (h)"
    chart.y_axis.title = "Heating Load (Wh)"

    # Reference X-axis data (Hour of the Year)
    hours_ref = Reference(workbook['Loads'], min_col=9, min_row=2, max_row=8761)  # Column 9: 'Hour of the Year'

    # Boiler Heating Data (Column 8: Boiler Heating)
    boiler_ref = Reference(workbook['Loads'], min_col=8, min_row=2, max_row=8761)
    boiler_series = Series(boiler_ref, hours_ref, title="Boiler Load (Wh)")
    boiler_series.marker.symbol = "circle"  # Set marker shape to circle
    boiler_series.graphicalProperties.line.noFill = True  # Remove connecting line
    boiler_series.marker.graphicalProperties.solidFill = "FF0000"  # Red color
    boiler_series.marker.graphicalProperties.line.solidFill = "000000"  # Black border
    boiler_series.marker.size = 5  # Set marker size

    # WSHP Heating Data (Column 6: WSHP Heating)
    wshp_ref = Reference(workbook['Loads'], min_col=6, min_row=2, max_row=8761)
    wshp_series = Series(wshp_ref, hours_ref, title="WSHP Load (Wh)")
    wshp_series.marker.symbol = "circle"  # Set marker shape to circle
    wshp_series.graphicalProperties.line.noFill = True  # Remove connecting line
    wshp_series.marker.graphicalProperties.solidFill = "00FF00"  # Green color
    wshp_series.marker.graphicalProperties.line.solidFill = "000000"  # Black border
    wshp_series.marker.size = 5  # Set marker size

    # ASHP Heating Data (Column 7: ASHP Heating)
    ashp_ref = Reference(workbook['Loads'], min_col=7, min_row=2, max_row=8761)
    ashp_series = Series(ashp_ref, hours_ref, title="ASHP Load (Wh)")
    ashp_series.marker.symbol = "circle"  # Set marker shape to circle
    ashp_series.graphicalProperties.line.noFill = True  # Remove connecting line
    ashp_series.marker.graphicalProperties.solidFill = "0000FF"  # Blue color
    ashp_series.marker.graphicalProperties.line.solidFill = "000000"  # Black border
    ashp_series.marker.size = 5  # Set marker size

    # Add the series to the chart
    chart.series.append(boiler_series)
    chart.series.append(wshp_series)
    chart.series.append(ashp_series)

    # Bring Boiler layer to the front by reordering
    chart.series = [boiler_series] + [series for series in chart.series if series != boiler_series]

    # Add the chart to the worksheet
    ws.add_chart(chart, "A1")  # Position the chart on the worksheet

    # Save the workbook
    workbook.save('Data/Output Data/ModelHomeData.xlsx')

# Reopen the workbook
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series

workbook = load_workbook('Data/Output Data/ModelHomeData.xlsx')
# Add the chart
create_scatter_chart(workbook)
# Save the workbook
workbook.save('Data/Output Data/ModelHomeData.xlsx')







# Use to verify correctness of SMARTS data - Create SSI file for May 1st at noon, and Solstice at 8pm and noon

# os.makedirs('Data/Output Data', exist_ok=True)
# df = pd.DataFrame({'Lambda (um)': lmbda, 'SSI (W*m^-2*nm^-1)': SSI[121, 11]})
# df.to_excel('Data/Output Data/SSI Verification/May 1st Noon SSI.xlsx', index=False)
#
# os.makedirs('Data/Output Data', exist_ok=True)
# df = pd.DataFrame({'Lambda (um)': lmbda, 'SSI (W*m^-2*nm^-1)': SSI[171, 11]}) # Remember 0 indexing
# df.to_excel('Data/Output Data/SSI Verification/Noon Solstice SSI.xlsx', index=False)
#
# os.makedirs('Data/Output Data', exist_ok=True)
# df = pd.DataFrame({'Lambda (um)': lmbda, 'SSI (W*m^-2*nm^-1)': SSI[171, 19]}) # Remember 0 indexing
# df.to_excel('Data/Output Data/SSI Verification/8pm Solstice SSI.xlsx', index=False)
